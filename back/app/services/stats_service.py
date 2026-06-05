from datetime import datetime
from typing import Optional
from io import BytesIO
import openpyxl
from app.repositories import stats_repo


def get_common(period_start: datetime, period_end: datetime,
               company_id: Optional[int] = None) -> dict:
    metrics = stats_repo.get_common_metrics(period_start, period_end, company_id)
    by_hours = stats_repo.get_tasks_by_hours(period_start, period_end, company_id)
    by_employees = stats_repo.get_tasks_by_employees(period_start, period_end, company_id)
    return {
        "period": {"from": period_start.date().isoformat(), "to": period_end.date().isoformat()},
        "tasks": metrics,
        "tasks_by_hours": by_hours,
        "tasks_by_employees": by_employees,
    }


def get_extended(period_start: datetime, period_end: datetime,
                 company_id: Optional[int] = None) -> dict:
    return {
        "by_unit_types": stats_repo.get_by_unit_types(period_start, period_end, company_id),
        "by_departments": stats_repo.get_by_departments(period_start, period_end, company_id),
        "by_unit_types_per_user": stats_repo.get_by_unit_types_per_user(period_start, period_end, company_id),
        "calendar": stats_repo.get_calendar(period_start, period_end, company_id),
    }


def get_responsibles(company_id: Optional[int] = None) -> list[dict]:
    return stats_repo.get_responsibles(company_id)


def get_user_tasks(user_id: int, period_start: datetime, period_end: datetime) -> dict:
    return stats_repo.get_user_tasks_detail(user_id, period_start, period_end)


def get_profile(user_id: int, period_start: datetime, period_end: datetime) -> dict:
    data = stats_repo.get_profile_stats(user_id, period_start, period_end)
    return {
        "period": {"from": period_start.date().isoformat(), "to": period_end.date().isoformat()},
        **data,
    }


def export_common_xlsx(period_start: datetime, period_end: datetime,
                       company_id: Optional[int] = None) -> BytesIO:
    data = get_common(period_start, period_end, company_id)
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Задачи за период"
    ws1.append(["Показатель", "Значение"])
    ws1.append(["Долг", data["tasks"]["debt"]])
    ws1.append(["Поступило", data["tasks"]["received"]])
    ws1.append(["Закрыто", data["tasks"]["closed"]])
    ws1.append(["Осталось", data["tasks"]["remaining"]])

    ws2 = wb.create_sheet("Задачи по часам")
    ws2.append(["Задача", "Суммарные часы"])
    for row in data["tasks_by_hours"]:
        ws2.append([row["name"], row["total_hours"]])

    ws3 = wb.create_sheet("По сотрудникам")
    ws3.append(["Сотрудник", "Задач", "Суммарные часы"])
    for row in data["tasks_by_employees"]:
        ws3.append([row["fio"], row["tasks_count"], row["total_hours"]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_extended_xlsx(period_start: datetime, period_end: datetime,
                         company_id: Optional[int] = None) -> BytesIO:
    data = get_extended(period_start, period_end, company_id)
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "По типам юнитов"
    ws1.append(["Тип юнита", "Суммарные часы", "Уникальных задач"])
    for row in data["by_unit_types"]:
        ws1.append([row["name"], row["total_hours"], row["tasks_count"]])

    ws2 = wb.create_sheet("По отделам")
    ws2.append(["Отдел", "Задач"])
    for row in data["by_departments"]:
        ws2.append([row["name"], row["tasks_count"]])

    ws3 = wb.create_sheet("По типам и сотрудникам")
    ws3.append(["Сотрудник", "Тип юнита", "Часы", "Задач"])
    for user_row in data["by_unit_types_per_user"]:
        for ut in user_row["unit_types"]:
            ws3.append([user_row["fio"], ut["name"], ut["hours"], ut["tasks_count"]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
